import pandas as pd
import csv
import re
import os
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def extract_homozygous_mutations(vcf_file, output_file):
    sample_list = [
        "ACT1_sorted", "ACT2_sorted", "ACT3_sorted", "ACT4_sorted", "ACT5_sorted", "B20_sorted", 
        "B21_sorted", "B22_sorted", "B23_sorted", "B24_sorted", "B25_sorted", "B26_sorted", 
        "B27_sorted", "B28_sorted", "B29_sorted", "B30_sorted", "B31_sorted", "B32_sorted", 
        "B33_sorted", "B34_sorted", "B35_sorted", "B36_sorted", "B37_sorted", "B38_sorted", 
        "B39_sorted", "B40_sorted", "B41_sorted", "B42_sorted", "B43_sorted", "B44_sorted", 
        "B45_sorted", "B46_sorted", "B47_sorted", "B48_sorted", "B49_sorted", "B50_sorted", 
        "B51_sorted", "B52_sorted"
    ]

    nonsynonymous_annotations = {"missense_variant"}

    with open(vcf_file, 'r') as vcf, open(output_file, 'w', newline='') as out_csv:
        reader = csv.reader(vcf, delimiter='\t')
        writer = csv.writer(out_csv, delimiter=';')

        # Écrire les en-têtes des colonnes
        writer.writerow(["POS", "REF", "ALT", "SAMPLES", "GENE", "ID", "MUTATION"])

        for line in reader:
            if line[0].startswith("#"):
                continue
            
            pos = line[1]
            ref = line[3]
            alt = line[4].split(",")  # Séparer les allèles alternatifs
            info = line[7]
            samples = line[9:]

            # Extraire le champ ANN pour obtenir les annotations
            annotation_field = [entry for entry in info.split(";") if entry.startswith("ANN=")]
            gene = "N/A"
            gene_id = "N/A"
            mutation = "N/A"
            annotation_type = "N/A"

            if annotation_field:
                annotations = annotation_field[0].split(",")
                for annotation in annotations:
                    fields = annotation.split("|")
                    if len(fields) >= 2:
                        annotation_type = fields[1]
                        if annotation_type in nonsynonymous_annotations:
                            gene = fields[3] if len(fields) >= 4 else "N/A"  # Nom du gène
                            gene_id = fields[4] if len(fields) >= 5 else "N/A"  # ID du gène (PF3D7_XXXXXXX)
                            mutation = fields[10] if len(fields) >= 11 else "N/A"  # Mutation protéique (p.L199V)
                            mutation = mutation.replace("p.", "")
                            break
                else:
                    continue

            homozygous_samples = []
            relevant_alt = set()
            for sample_name, sample_info in zip(sample_list, samples):
                genotype = sample_info.split(":")[0]
                if genotype == "1/1":
                    homozygous_samples.append(sample_name)
                    relevant_alt.add(alt[0])
                elif genotype == "2/2" and len(alt) > 1:
                    homozygous_samples.append(sample_name)
                    relevant_alt.add(alt[1])
                elif genotype == "3/3" and len(alt) > 2:
                    homozygous_samples.append(sample_name)
                    relevant_alt.add(alt[2])
                elif genotype == "4/4" and len(alt) > 3:
                    homozygous_samples.append(sample_name)
                    relevant_alt.add(alt[3])
            
            if homozygous_samples:
                writer.writerow([pos, ref, ",".join(relevant_alt), ",".join(homozygous_samples), gene, gene_id, mutation])


def convert_mutation_file(input_file):
    """
    Convert mutation codes in a CSV file from three-letter to one-letter notation.

    Parameters:
    - input_file: Path to the input CSV file.
    - column_name: The name of the column containing mutations to be converted (default is "MUTATION").
    """
    amino_acid_map = {
        "Ala": "A", "Arg": "R", "Asn": "N", "Asp": "D",
        "Cys": "C", "Gln": "Q", "Glu": "E", "Gly": "G",
        "His": "H", "Ile": "I", "Leu": "L", "Lys": "K",
        "Met": "M", "Phe": "F", "Pro": "P", "Ser": "S",
        "Thr": "T", "Trp": "W", "Tyr": "Y", "Val": "V"
    }
    column_name="MUTATION"

    def convert_mutation(mutation):
        # Séparer la mutation en trois parties : premier acide aminé, position, second acide aminé
        match = re.match(r"([A-Za-z]{3})(\d+)([A-Za-z]{3})", mutation)
        if match:
            aa1, pos, aa2 = match.groups()
            # Convertir les codes 3 lettres en 1 lettre
            aa1_one = amino_acid_map.get(aa1, aa1)
            aa2_one = amino_acid_map.get(aa2, aa2)
            return f"{aa1_one}{pos}{aa2_one}"
        return mutation  # Retourner inchangé si le format est incorrect

    try:
        df = pd.read_csv(input_file, sep=';', encoding='utf-8')
    except UnicodeDecodeError:
        df = pd.read_csv(input_file, sep=';', encoding='latin1') 

    # Vérifier si la colonne existe
    if column_name not in df.columns:
        raise ValueError(f"La colonne '{column_name}' n'existe pas dans le fichier.")

    # Appliquer la conversion à la colonne spécifiée
    df[column_name] = df[column_name].apply(convert_mutation)

    # Sauvegarder le fichier modifié
    df.to_csv(input_file, index=False, sep=';', encoding='utf-8')
    print(f"Fichier traité et sauvegardé : {input_file}")


def calculate_occurrences(mut_file, haplotype_folder):
    """
    Ajoute les colonnes AF (somme des colonnes commençant par "AF-") et Occurrences 
    (nombre de fois où une mutation apparaît dans ns_changes) au fichier MRP2.
    
    Parameters:
    - mut_file: chemin vers le fichier MRP2 (CSV).
    - classeur1_file: chemin vers le fichier Classeur1 (CSV).
    - output_file: chemin où sauvegarder le fichier avec les colonnes ajoutées.
    - mutation_column: nom de la colonne contenant les mutations dans MRP2.
    - ns_changes_column: nom de la colonne contenant les mutations dans Classeur1.
    - af_prefix: préfixe des colonnes à sommer (exemple : "AF").
    """
    # Charger les fichiers CSV
    mut_df = pd.read_csv(mut_file, sep=";")

    # Ajouter des colonnes pour le total AF
    mut_df["AF"] = 0
    mut_df["AS"] = 0
    mut_df["SA"] = 0
    mut_df["OC"] = 0

    dict_mutation_id = dict(zip(mut_df["MUTATION"], mut_df["ID"])) 

    for mutation, gene_id in dict_mutation_id.items():
        filename_pattern = f"pf-haploatlas-{gene_id}_population_summary.csv"
        for filename in os.listdir(haplotype_folder):
            if filename.startswith(filename_pattern):
                haplotype_file = os.path.join(haplotype_folder, filename)
                haplotype_df = pd.read_csv(haplotype_file, sep=",")  # Adapter le séparateur si nécessaire

                # Sélectionner les colonnes spécifiques
                af_columns = [col for col in haplotype_df.columns if col.startswith("AF")]
                as_columns = [col for col in haplotype_df.columns if col.startswith("AS")]
                sa_columns = [col for col in haplotype_df.columns if col.startswith("SA")]
                oc_columns = [col for col in haplotype_df.columns if col.startswith("OC")]

                # Vérifier si la colonne "ns_changes" existe
                if "ns_changes" not in haplotype_df.columns:
                    print(f"Avertissement : La colonne 'ns_changes' est absente dans {filename}")
                    continue

                # Filtrer les lignes contenant la mutation
                filtered_rows = haplotype_df[haplotype_df["ns_changes"].str.contains(mutation, na=False)]

                # Calculer la somme des valeurs dans les colonnes d'intérêt
                total_af = filtered_rows[af_columns].sum().sum() if not filtered_rows.empty else 0
                total_as = filtered_rows[as_columns].sum().sum() if not filtered_rows.empty else 0
                total_sa = filtered_rows[sa_columns].sum().sum() if not filtered_rows.empty else 0
                total_oc = filtered_rows[oc_columns].sum().sum() if not filtered_rows.empty else 0

                # Mettre à jour les valeurs **spécifiques** à la mutation
                mut_df.loc[mut_df["MUTATION"] == mutation, "AF"] = total_af
                mut_df.loc[mut_df["MUTATION"] == mutation, "AS"] = total_as
                mut_df.loc[mut_df["MUTATION"] == mutation, "SA"] = total_sa
                mut_df.loc[mut_df["MUTATION"] == mutation, "OC"] = total_oc

                print(f"Mutation {mutation} ({gene_id}) -> AF={total_af}, AS={total_as}, SA={total_sa}, OC={total_oc}")

    mut_df.to_csv(mut_file, sep=";", index=False)
    print(f"Fichier mis à jour et sauvegardé sous : {mut_file}")


def colour():
    csv_file = "C:\\Users\\Utilisateur\\Downloads\\stage_noe\\vcf_filter\\filtered_mutations.csv"
    output_excel = "C:\\Users\\Utilisateur\\Downloads\\stage_noe\\vcf_filter\\filtered_mutations_coloured.xlsx"
    highlight_values = ["B20", "B21", "B22", "B23", "B24", "B25", "B26", "B27", "B28", "B29", "B30", "B31"]
    


    df = pd.read_csv(csv_file, sep=";")
    df.to_excel(output_excel, index=False, engine='openpyxl')
    wb = load_workbook(output_excel)
    ws = wb.active
    green_fill = PatternFill(start_color="01796F", end_color="01796F", fill_type="solid")
    sample_col_index = None
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == "SAMPLES":
            sample_col_index = col[0].column
            break

    if sample_col_index:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=sample_col_index, max_col=sample_col_index):
            cell = row[0]
            if any(value in str(cell.value).replace("_sorted", "").split(",") for value in highlight_values):  # Vérification améliorée
                for c in ws[row[0].row]:
                    c.fill = green_fill  

    wb.save(output_excel)
    print(f"Fichier coloré sauvegardé sous : {output_excel}")


def filter(input_file, output_file):
    # Lire le CSV
    df = pd.read_csv(input_file, sep=";")  # Vérifie que le séparateur est correct

    # Filtrer les lignes où AF == 0
    df_filtered = df[df["AF"] == 0]

    # Sauvegarder le fichier filtré
    df_filtered.to_csv(output_file, sep=";", index=False)

    print(f"Fichier filtré enregistré sous : {output_file}")


vcf_file = "C:\\Users\\User\\Downloads\\stage_noe\\vcf_filter\\final_noe.vcf"
output_file = "C:\\Users\\User\\Downloads\\stage_noe\\vcf_filter\\filtered_mutations.csv"
haplotype_folder = "C:\\Users\\User\\Downloads\\stage_noe\\vcf_filter\\malariagen_noe"

#extract_homozygous_mutations(vcf_file, output_file)
#convert_mutation_file(output_file)
#calculate_occurrences(output_file, haplotype_folder)
#colour()
#filter(output_file, output_file="C:\\Users\\Utilisateur\\Downloads\\stage_noe\\vcf_filter\\filtered_mutations_0AF.csv")
